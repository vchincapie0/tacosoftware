# Fecha de Creación: 20/02/2024
# Autor: Vivian Carolina Hincapie Escobar
# Última modificación: 15/05/2024

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import (
    ListView, 
    CreateView, 
    UpdateView, 
    DeleteView)
from django.urls import reverse_lazy

#Importacion de modelos y formularios
from .models import Pedidos, PedidosAudit
from applications.materiaprima.models import MateriaPrima
from applications.insumos.models import Insumos
from applications.proveedores.models import Proveedores

from .forms import (
    PedidosCreateForm, 
    PedidosUpdateForm, 
    PedidosAddMpCreateFrom, 
    PedidosAddInsumosCreateFrom,
    PedidosAddProveedorCreateFrom,
    PedidosAuditFilterForm

)

class PedidosListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los Pedidos'''
    model = Pedidos
    template_name = "pedidos/list_pedidos.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=5
    context_object_name = 'pedidos'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la pablabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = Pedidos.objects.filter(
            ref_pedido__icontains = palabra_clave,
            deleted=False  # Solo mostrará pedidos activos
        )
        return lista

class PedidosCreateView(LoginRequiredMixin, CreateView):

    '''Clase donde se crea un nuevo pedido recibido'''
    model = Pedidos
    template_name = "pedidos/add_pedidos.html"
    login_url=reverse_lazy('home_app:home')
    #Campos que se van a mostrar en el formulario
    form_class = PedidosCreateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('pedidos_app:list_pedidos')

    def form_valid(self, form):
        '''Funcion para validar el formulario de crear pedido'''
        referencia = form.cleaned_data['ref_pedido']
        user = self.request.user
            # Guarda el formulario sin commit para asignar manualmente el usuario
        pedido = form.save(commit=False)
            # Asigna el usuario al campo pedi_user
        pedido.pedi_user = user
            # Ahora sí, guarda el pedido en la base de datos
        pedido.save()

        # Agregar un mensaje de éxito con el numero de referencia
        messages.success(self.request, f'¡El pedido {referencia} se ha agregado correctamente!')

        return super(PedidosCreateView, self).form_valid(form)

class PedidosAddMpCreateView(LoginRequiredMixin,CreateView):
    '''Clase para crear una materia prima nueva dentro del formulario de agregar pedidos'''
    model = MateriaPrima
    template_name = "pedidos/add_mp_pedidos.html"
    login_url=reverse_lazy('home_app:home')
    #Campos que se van a mostrar en el formulario
    form_class = PedidosAddMpCreateFrom
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('pedidos_app:add_pedidos') 

class PedidosAddInsumosCreateView(LoginRequiredMixin,CreateView):
    '''Clase para crear un implemento de trabajo nuevo dentro del formulario de agregar pedidos'''
    model = Insumos
    template_name = "pedidos/add_it_pedidos.html"
    login_url=reverse_lazy('home_app:home')
    #Campos que se van a mostrar en el formulario
    form_class = PedidosAddInsumosCreateFrom
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('pedidos_app:add_pedidos')

class PedidosAddProveedoresCreateView(LoginRequiredMixin,CreateView):

    '''Clase para crear un proveedor nuevo dentro del formulario de agregar pedidos'''
    model = Proveedores
    template_name = "pedidos/add_prov_pedidos.html"
    login_url=reverse_lazy('home_app:home')
    #Campos que se van a mostrar en el formulario
    form_class = PedidosAddProveedorCreateFrom
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('pedidos_app:add_pedidos') 

class PedidosUpdateView(LoginRequiredMixin, UpdateView):

    '''Vista para actualizar los datos de proveedores'''
    model = Pedidos
    template_name = "pedidos/edit_pedidos.html"
    login_url=reverse_lazy('users_app:login')
    form_class=PedidosUpdateForm
    success_url= reverse_lazy('pedidos_app:list_pedidos')

    def form_valid(self, form):
        '''función para validar formulario de update pedido y mandar mensaje de confirmación'''
        referencia = form.cleaned_data['ref_pedido']

         # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡El pedido {referencia} se ha actualizado correctamente!')

        return super(PedidosUpdateView, self).form_valid(form)

class PedidosDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Pedidos'''
    model = Pedidos
    template_name = "pedidos/delete_pedidos.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('pedidos_app:list_pedidos')

class PedidosAuditListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de la Auditoria de Pedidos'''
    model = PedidosAudit
    template_name = "administrador/auditorias/pedidosaudit.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'pedidos'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = PedidosAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            pedido = form.cleaned_data.get('pedido')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if pedido:
                queryset = queryset.filter(pedido=pedido)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = PedidosAuditFilterForm(self.request.GET)
        return context
    
def export_pedidos_to_excel(request):
    '''Vista que se encarga de la descarga de datos de la tabla pedidos en formato excel'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de pedidos que quieres exportar
    pedidos = Pedidos.objects.filter(deleted=False)  # Filtrar pedidos activos

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Transparente
    title_alignment = Alignment(horizontal='center')
    
    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])  # Agregar texto del título
    worksheet.merge_cells('A1:H1')  # Combinar celdas para el título
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])
    worksheet.append(['Registros','Pedidos Registrados'])

    # Agregar espacio en blanco entre la información adicional y los encabezados
    worksheet.append([])  # Agregar una fila vacía

    # Agregar encabezados a la siguiente fila
    headers = ['Referencia', 'Responsable', 'Fecha de Recibido', 'Estado', 'Comprobante de Pago', 'Proveedor', 'Materia Prima', 'Implementos de Trabajo']
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gris claro
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de pedidos a las siguientes filas
    for pedido in pedidos:
        # Obtener los valores necesarios de los modelos relacionados
        proveedor_nombre = pedido.pedi_proveedor.prov_nombre if pedido.pedi_proveedor else ''
        user_nombre = pedido.pedi_user.name if pedido.pedi_user else ''
            
        # Obtener nombres de materias primas como una lista de cadenas
        materia_prima_list = [str(materia.mp_nombre) for materia in pedido.pedi_materiaprima.all()] if pedido.pedi_materiaprima.exists() else []
        materia_prima_str = ', '.join(materia_prima_list)  # Convertir lista a una cadena separada por comas

            
        # Obtener nombres de implementos de trabajo como una lista de cadenas
        implementos_trabajo_list = [implemento.it_nombre.it_nombre for implemento in pedido.pedi_insumos.all()] if pedido.pedi_insumos.exists() else []
        implementos_trabajo_str = ', '.join(implementos_trabajo_list)  # Convertir lista a una cadena separada por comas 
        # Construir la fila de datos para el pedido
        data_row = [
            pedido.ref_pedido,
            user_nombre,
            pedido.pedi_fecha,
            pedido.get_pedi_estado_display(),  # Mostrar el nombre del estado en lugar del código
            pedido.pedi_comprobatePago,
            proveedor_nombre,
            materia_prima_str,
            implementos_trabajo_str,
        ]

        # Agregar la fila de datos a la hoja de trabajo
        worksheet.append(data_row)


    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

def export_pedidos_to_csv(request):
    '''Vista que se encarga de la descarga de datos de la tabla pedidos en formato csv'''
    pedidos = Pedidos.objects.filter(deleted=False)  # Obtener datos de pedidos
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=pedidos.csv'

    writer = csv.writer(response)
    writer.writerow(['Referencia', 'Responsable', 'Fecha de Recibido', 'Estado', 'Comprobante de Pago', 'Proveedor', 'Materia Prima', 'Implementos de Trabajo'])  # Encabezados de columnas

    # Agregar datos de pedidos a las siguientes filas
    for pedido in pedidos:
        # Obtener los valores necesarios de los modelos relacionados
        proveedor_nombre = pedido.pedi_proveedor.prov_nombre if pedido.pedi_proveedor else ''
        user_nombre = pedido.pedi_user.name if pedido.pedi_user else ''
            
        # Obtener nombres de materias primas como una lista de cadenas
        materia_prima_list = [str(materia.mp_nombre) for materia in pedido.pedi_materiaprima.all()] if pedido.pedi_materiaprima.exists() else []
        materia_prima_str = ', '.join(materia_prima_list)  # Convertir lista a una cadena separada por comas

            
        # Obtener nombres de implementos de trabajo como una lista de cadenas
        implementos_trabajo_list = [implemento.it_nombre.it_nombre for implemento in pedido.pedi_insumos.all()] if pedido.pedi_insumos.exists() else []
        implementos_trabajo_str = ', '.join(implementos_trabajo_list)  # Convertir lista a una cadena separada por comas 
        # Construir la fila de datos para el pedido
        writer.writerow ([
            pedido.ref_pedido,
            user_nombre,
            pedido.pedi_fecha,
            pedido.get_pedi_estado_display(),  # Mostrar el nombre del estado en lugar del código
            pedido.pedi_comprobatePago,
            proveedor_nombre,
            materia_prima_str,
            implementos_trabajo_str,
        ])

    return response
