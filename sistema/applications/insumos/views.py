from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.contrib import messages
from django.shortcuts import render
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
#Importacion de modelos y formularios
from .models import Insumos,InsumosGenerico,InsumosAudit
from .forms import  InsumosUpdateForm,InsumosGenericoForm,InsumosGenericoUpdateForm,InsumosAuditFilterForm

# Create your views here.

class InsumosGenericoListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los Insumos Genericos'''
    model = InsumosGenerico
    template_name = "administrador/genericas/list_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'insumos'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la pablabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = InsumosGenerico.objects.filter(
           it_nombre__icontains = palabra_clave
        )
        return lista
    
class InsumosGenericoCreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea una nueva materia prima'''
    model = InsumosGenerico
    template_name = "administrador/genericas/add_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class=InsumosGenericoForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('insumos_app:list_insumos_generico')

class InsumosGenericoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de user'''
    model = InsumosGenerico
    template_name = "administrador/genericas/update_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    form_class=InsumosGenericoUpdateForm
    success_url= reverse_lazy('insumos_app:list_insumos')

class InsumosGenericoDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Insumos'''
    model = InsumosGenerico
    template_name = "administrador/genericas/delete_insumos_generico.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('insumos_app:list_insumos_generico')

class InsumosListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los Insumos'''
    model = Insumos
    template_name = "insumos/list_insumos.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'insumos'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la palabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = Insumos.objects.filter(
           it_nombre__it_nombre__icontains = palabra_clave
        )
        return lista
    
class InsumosUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de user'''
    model = Insumos
    template_name = "insumos/update_insumos.html"
    login_url=reverse_lazy('users_app:login')
    form_class=InsumosUpdateForm
    success_url= reverse_lazy('insumos_app:list_insumos')

    def form_valid(self, form):
        #Obtener los datos del fomulario
        nombre = form.cleaned_data['it_nombre']

        # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡El insumo {nombre} se ha actualizado correctamente!')

        return super(InsumosUpdateView, self).form_valid(form)
    
class InsumosAuditListView(LoginRequiredMixin, ListView):
    model= InsumosAudit
    template_name='administrador/auditorias/insumosaudit.html'
    paginate_by=10
    context_object_name='auditoria'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = InsumosAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            insumos = form.cleaned_data.get('insumos')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if insumos:
                queryset = queryset.filter(insumos=insumos)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = InsumosAuditFilterForm(self.request.GET)
        return context

def export_insumos_to_excel(request):
    '''Vista para exportar datos de tabla insumos en formato excel'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de insumos que quieres exportar
    insumos = Insumos.objects.filter(deleted=False)  # Filtrar insumos activos

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Insumos'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Transparente
    title_alignment = Alignment(horizontal='center')
    
    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])  # Agregar texto del título
    worksheet.merge_cells('A1:D1')  # Combinar celdas para el título
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])

    # Agregar espacio en blanco entre la información adicional y los encabezados
    worksheet.append([])  # Agregar una fila vacía

    # Agregar encabezados a la siguiente fila
    headers = ['Insumo', 'Cantidad', 'Fecha de Entrega','Estado']
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gris claro
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de implementos a las siguientes filas
    for insumo in insumos:
        data_row = [
            insumo.it_nombre.it_nombre,  # Access related field's name
            insumo.it_cantidad,
            insumo.it_fechaEntrega.strftime("%Y-%m-%d"),  # Format date
            insumo.get_it_estado_display()  # Get display value of choice field
        ]
        worksheet.append(data_row)

    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=insumos.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

def export_insumos_to_csv(request):
    '''Vista para exportar datos de tabla insumos en formato CSV'''
    insumos = Insumos.objects.filter(deleted=False)  # Obtener datos de proveedores
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=insumos.csv'

    writer = csv.writer(response)
    writer.writerow(['Insumo', 'Cantidad', 'Fecha de Entrega','Estado'])  # Encabezados de columnas

    for insumo in insumos:
        writer.writerow([
            insumo.it_nombre.it_nombre,  # Access related field's name
            insumo.it_cantidad,
            insumo.it_fechaEntrega.strftime("%Y-%m-%d"),  # Format date
            insumo.get_it_estado_display()  # Get display value of choice field
        ])


    return response