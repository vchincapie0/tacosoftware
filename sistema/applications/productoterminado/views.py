from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Count
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
#Importacion de modelos y formularios
from .models import (
    ProductoTerminado,
    ExistenciaPT,
    CaracteristicasOrganolepticasPT,
    EmpaqueProductoTerminado,
    Vacio,
    ProductoTerminadoGenerico,
    ProductoTerminadoGenerico,
    ProductoTerminadoAudit
    
)
from .forms import (
    ProductoTerminadoForm,
    CaracteristicasOrganolepticasPTForm,
    EmpaqueProductoTerminadoForm,
    VacioForm,
    CaracteristicasPTUpdateForm,
    EmpaqueUpdateForm,
    VacioUpdateForm,
    ProductoTerminadoGenericoForm,
    ProductoTerminadoGenericoFilterForm,
    ProductoAuditFilterForm
)

class ProduListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/list_produ.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'productoterminado'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la pablabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = ProductoTerminado.objects.filter(
            pt_nombre__pt_nombre__icontains = palabra_clave
        )
        return lista

class ProduUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/update_produ.html"
    login_url=reverse_lazy('users_app:login')
    form_class=ProductoTerminadoForm
    success_url= reverse_lazy('produ_app:list_produ')
    

    def form_valid(self, form):
        #Obtener los datos del fomulario
        pt_nombre = form.cleaned_data['pt_nombre']
        pt_fecha = form.cleaned_data['pt_fechapreparacion']


        # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡El producto {pt_nombre} de la fecha de preparación {pt_fecha} se ha actualizado correctamente!')

        return super(ProduUpdateView, self).form_valid(form)

class CaracteristicasProductoTerminadoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para la edición de las caracteristicas organolepticas de producto terminado'''
    model = CaracteristicasOrganolepticasPT
    template_name = "productoterminado/updatecaracteristicas_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = CaracteristicasPTUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')

    def form_valid(self, form):
        #Obtener los datos del fomulario
        pt_lote = form.cleaned_data['pt_lote']
        producto_nombre = pt_lote.pt_nombre.pt_nombre
               
        # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡Las características de {producto_nombre} se ha actualizado correctamente!')

        return super(CaracteristicasProductoTerminadoUpdateView, self).form_valid(form)

class ProductoTerminadoDetailView(LoginRequiredMixin, DetailView):
    
    '''Vista donde se muestran los detalles de producto terminado'''
    model = ProductoTerminado
    template_name = "productoterminado/detail_PT.html"
    login_url=reverse_lazy('users_app:login')
    context_object_name = 'productoterminado'

    def get_object(self, queryset=None):
        lote = self.kwargs.get('pt_lote')
        return self.model.objects.get(pt_lote=lote)
    
class EmpaqueProductoTerminadoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vists para la edición del empaque producto terminado'''
    model = EmpaqueProductoTerminado
    template_name = "productoterminado/empaqueupdate_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = EmpaqueUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')
    
    def form_valid(self, form):
        '''funcion para automatizar el campo '''
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        empaque = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        empaque.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        empaque.save()
        return super().form_valid(form)    

    def form_valid(self, form):
        #Obtener los datos del fomulario
        pt_lote = form.cleaned_data['pt_lote']
        
        # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡El empacado de {pt_lote} se ha actualizado correctamente!')

        return super(EmpaqueProductoTerminadoUpdateView, self).form_valid(form)
    
class VacioProductoTerminadoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vists para la edición del vacio producto terminado'''
    model = Vacio
    template_name = "productoterminado/vacioupdate_pt.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = VacioUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_produ')
    
    def form_valid(self, form):
        '''funcion para automatizar el campo '''
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        Vacio = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        Vacio.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        Vacio.save()
        return super().form_valid(form)   
    def form_valid(self, form):
        #Obtener los datos del fomulario
        pt_lote = form.cleaned_data['pt_lote']
        
        # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡El empacado al vacio de {pt_lote} se ha actualizado correctamente!')

        return super(VacioProductoTerminadoUpdateView, self).form_valid(form)

from django.db.models import Count

class ProductoTerminadoGenericoListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de las materias primas'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/productoterminado/list_pt_generico.html"
    login_url = reverse_lazy('users_app:login')
    paginate_by = 10
    context_object_name = 'producto'

    def get_queryset(self):
        queryset = super().get_queryset()
        pt_nombre = self.request.GET.get('pt_nombre', None)
        materia_prima_ids = self.request.GET.getlist('materiaPrimaUsada')

        if pt_nombre:
            queryset = queryset.filter(pt_nombre__icontains=pt_nombre)

        if materia_prima_ids:
            # Filtra productos que contienen al menos una de las materias primas seleccionadas
            queryset = queryset.filter(materiaPrimaUsada__id__in=materia_prima_ids)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProductoTerminadoGenericoFilterForm(self.request.GET or None)
        return context

    
class ProductoTerminadoGenericoCreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea un nuevo Producto terminado'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/productoterminado/add_pt_generico.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = ProductoTerminadoGenericoForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('produ_app:list_pt_generico')

class ProductoTerminadoGenericoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos  Producto terminado'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/productoterminado/update_pt_generico.html"
    login_url=reverse_lazy('users_app:login')
    form_class=ProductoTerminadoGenericoForm
    success_url= reverse_lazy('produ_app:list_pt_generico')

class ProductoTerminadoGenericoDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Producto terminado'''
    model = ProductoTerminadoGenerico
    template_name = "administrador/genericas/productoterminado/delete_pt_generico.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('produ_app:list_pt_generico')

class ProductoAuditListView(LoginRequiredMixin, ListView):
    model= ProductoTerminadoAudit
    template_name='administrador/auditorias/productoaudit.html'
    paginate_by=10
    context_object_name='auditoria'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = ProductoAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            productoterminado = form.cleaned_data.get('productoterminado')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if productoterminado:
                queryset = queryset.filter(productoterminado=productoterminado)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProductoAuditFilterForm(self.request.GET)
        return context
    
def export_productoterminado_to_excel(request):
    '''Vista para exportar datos de tabla producto terminado en formato excel'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de producto terminado que quieres exportar
    productoterminado = ProductoTerminado.objects.all()

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Producto Terminado'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    title_alignment = Alignment(horizontal='center')

    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])
    worksheet.merge_cells('A1:O1')
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])

    #blanco entre la información adicional y los encabezados
    worksheet.append([])

    # Agregar fila de título de características organolépticas
    worksheet.merge_cells('G5:K5')  # Fusionar celdas para el título
    worksheet['G5'] = 'CARACTERISTICAS ORGANOLEPTICAS'  # Escribir el título en la celda fusionada

    # Establecer estilos para el título de características organolépticas
    title_font = Font(bold=True, color="000000")  # Letra negra
    title_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro
    title_alignment = Alignment(horizontal='center')

    # Aplicar estilos al título
    title_cell = worksheet['G5']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    

    # Agregar encabezados a la siguiente fila
    headers = [
        'Lote', 'Producto', 'Cantidad', 'Fecha de preparación', 'Fecha de vencimiento',
        'Observaciones', 'Olor', 'Sabor', 'Textura', 'Color', 'Estado',
        'Peso Empaque (Kg)', 'Cantidad Bolsas', 
        'Bolsas Rechazadas', 'Bolsas Liberadas'
    ]
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de productos terminados a las siguientes filas
    for producto in productoterminado:
        # Obtener las características organolépticas
        try:
            caracteristicas = CaracteristicasOrganolepticasPT.objects.get(pt_lote=producto)
        except CaracteristicasOrganolepticasPT.DoesNotExist:
            caracteristicas = None

        # Obtener los datos de empaque
        try:
            empaque = EmpaqueProductoTerminado.objects.get(pt_lote=producto)
        except EmpaqueProductoTerminado.DoesNotExist:
            empaque = None

        # Obtener los datos de vacío
        try:
            vacio = Vacio.objects.get(pt_lote=producto)
        except Vacio.DoesNotExist:
            vacio = None

        data_row = [
            producto.pt_lote,
            producto.pt_nombre.pt_nombre,
            producto.pt_cantidad,
            producto.pt_fechapreparacion.strftime("%Y-%m-%d"),
            producto.pt_fechavencimiento.strftime("%Y-%m-%d"),
            caracteristicas.observaciones if caracteristicas else '',
            'Sí' if caracteristicas and caracteristicas.olor else 'No',
            'Sí' if caracteristicas and caracteristicas.sabor else 'No',
            'Sí' if caracteristicas and caracteristicas.textura else 'No',
            'Sí' if caracteristicas and caracteristicas.color else 'No',
            dict(CaracteristicasOrganolepticasPT.ESTADO_CHOICES).get(caracteristicas.estado, '') if caracteristicas else '',
            empaque.emp_pesoKg if empaque else '',
            empaque.emp_cantidadBolsas if empaque else '',
            vacio.cantidad_bolsas_rechazadas if vacio else '',
            vacio.cantidad_bolsas_liberadas if vacio else ''
        ]
        worksheet.append(data_row)

    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productoterminado.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

def export_productoterminado_to_csv(request):
    '''Vista para exportar datos de tabla producto terminado en formato CSV'''
    productoterminado = ProductoTerminado.objects.all()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=productoterminado.csv'

    writer = csv.writer(response)
    headers = [
        'Lote', 'Producto', 'Cantidad', 'Fecha de preparación', 'Fecha de vencimiento',
        'Observaciones', 'Olor', 'Sabor', 'Textura', 'Color', 'Estado',
        'Peso Empaque (Kg)', 'Cantidad Bolsas', 
        'Bolsas Rechazadas', 'Bolsas Liberadas'
    ]
    writer.writerow(headers)

    for producto in productoterminado:
        # Obtener las características organolépticas
        try:
            caracteristicas = CaracteristicasOrganolepticasPT.objects.get(pt_lote=producto)
        except CaracteristicasOrganolepticasPT.DoesNotExist:
            caracteristicas = None

        # Obtener los datos de empaque
        try:
            empaque = EmpaqueProductoTerminado.objects.get(pt_lote=producto)
        except EmpaqueProductoTerminado.DoesNotExist:
            empaque = None

        # Obtener los datos de vacío
        try:
            vacio = Vacio.objects.get(pt_lote=producto)
        except Vacio.DoesNotExist:
            vacio = None

        writer.writerow([
            producto.pt_lote,
            producto.pt_nombre.pt_nombre,
            producto.pt_cantidad,
            producto.pt_fechapreparacion.strftime("%Y-%m-%d"),
            producto.pt_fechavencimiento.strftime("%Y-%m-%d"),
            caracteristicas.observaciones if caracteristicas else '',
            'Sí' if caracteristicas and caracteristicas.olor else 'No',
            'Sí' if caracteristicas and caracteristicas.sabor else 'No',
            'Sí' if caracteristicas and caracteristicas.textura else 'No',
            'Sí' if caracteristicas and caracteristicas.color else 'No',
            dict(CaracteristicasOrganolepticasPT.ESTADO_CHOICES).get(caracteristicas.estado, '') if caracteristicas else '',
            empaque.emp_pesoKg if empaque else '',
            empaque.emp_cantidadBolsas if empaque else '',
            vacio.cantidad_bolsas_rechazadas if vacio else '',
            vacio.cantidad_bolsas_liberadas])
        return response