from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView,CreateView,DetailView, UpdateView, TemplateView, DeleteView
from django.utils import timezone
from django.urls import reverse_lazy
from django.shortcuts import render

#Importacion modelos y formularios
from applications.pedidos.models import Pedidos
from .models import (
    MateriaPrima,Desinfeccion,
    CaracteristicasOrganolepticas,
    MateriaPrimaGenerica,
    DesinfectanteGenerico,
    MateriaPrimaAudit
    )
from .forms import (
    
    MateriaPrimaUpdateForm,
    CaracteristicasMPForm,
    CaracteristicasMPUpdateForm,
    DesinfeccionMPForm,
    DesinfeccionMPUpdateForm,
    MateriaPrimaGenericaForm,
    MateriaPrimaGenericaUpdateForm,
    DesinfectanteGenericoForm,
    MateriaAuditFilterForm
)


# Create your views here.
class MateriaPrimaGenericaListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de las materias primas'''
    model = MateriaPrimaGenerica
    template_name = "administrador/genericas/lista_mp_generica.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'materiaprima'

    def get_queryset(self):
        palabra_clave = self.request.GET.get("kword", '')
        
        # Filtrar por nombre específico de la materia prima
        queryset = MateriaPrimaGenerica.objects.filter(
            mp_nombre__icontains=palabra_clave
        )
        
        return queryset

class MateriaPrimaGenericaCreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea una nueva materia prima'''
    model = MateriaPrimaGenerica
    template_name = "administrador/genericas/add_mp_generica.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = MateriaPrimaGenericaForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('mp_app:listaGenerica_mp')

class MateriaPrimaGenericaUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de materia prima generica'''
    model = MateriaPrimaGenerica
    template_name = "administrador/genericas/update_mp_generica.html"
    login_url=reverse_lazy('users_app:login')
    form_class=MateriaPrimaGenericaUpdateForm
    success_url= reverse_lazy('mp_app:listaGenerica_mp')

class MateriaPrimaGenericaDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Implenentos de Trabajo'''
    model = MateriaPrimaGenerica
    template_name = "administrador/genericas/delete_mp_generica.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('mp_app:listaGenerica_mp')

class MateriaPrimaListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de las materias primas'''
    model = MateriaPrima
    template_name = "materiaprima/lista_mp.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'materiaprima'

    def get_queryset(self):
        palabra_clave = self.request.GET.get("kword", '')

        # Obtiene todas las materias primas
        queryset = MateriaPrima.objects.all()

        # Filtra pedidos rechazados
        pedidos_rechazados_ids = Pedidos.objects.filter(pedi_estado='2').values_list('id', flat=True)

        # Filtra las materias primas excluyendo aquellas asociadas a pedidos rechazados
        queryset = queryset.exclude(pedidos__id__in=pedidos_rechazados_ids)

        # Filtra por nombre específico de la materia prima si se proporciona una palabra clave
        if palabra_clave:
            queryset = queryset.filter(mp_nombre__icontains=palabra_clave)

        return queryset

   
class MateriaPrimaUpdateView(LoginRequiredMixin, UpdateView):
    '''Clase donde se modifica la materia prima registrada'''
    model = MateriaPrima
    template_name = "materiaprima/mp_update.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = MateriaPrimaUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('mp_app:lista_mp') 

    def form_valid(self, form):
        #Obtener los datos del fomulario
        mp_nombre = form.cleaned_data['mp_nombre']

        # Agregar un mensaje de éxito con el nombre de la materia prima
        messages.success(self.request, f'¡La materia prima {mp_nombre} se ha actualizado correctamente!')

        return super(MateriaPrimaUpdateView, self).form_valid(form)


class CaracteristicasMateriaPrimaCreateView(LoginRequiredMixin, CreateView):
    '''Vista para la creacion de las caracteristicas organolepticas de la materia prima'''
    model = CaracteristicasOrganolepticas
    template_name = "materiaprima/caracteristicas_mp.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = CaracteristicasMPForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('mp_app:lista_mp')

    def form_valid(self, form):
        #Obtener los datos del fomulario
        mp_lote = form.cleaned_data['mp_lote']

        # Agregar un mensaje de éxito con el nombre de la materia prima
        messages.success(self.request, f'¡Las características {mp_lote} se han guardado correctamente!')

        return super(CaracteristicasMateriaPrimaCreateView, self).form_valid(form)
    
class CaracteristicasMateriaPrimaUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para la edicion de las caracteristicas organolepticas de la materia prima'''
    model = CaracteristicasOrganolepticas
    template_name = "materiaprima/updateCaracteristicas_mp.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = CaracteristicasMPUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('mp_app:lista_mp')
   
    def form_valid(self, form):
        #Obtener los datos del fomulario
        mp_lote = form.cleaned_data['mp_lote']

        # Agregar un mensaje de éxito con el nombre de la materia prima
        messages.success(self.request, f'¡Las características {mp_lote} se han actulizado correctamente!')

        return super(CaracteristicasMateriaPrimaUpdateView, self).form_valid(form)
    
    
class DesinfectanteGenericoListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los Implementos de trabajo'''
    model = DesinfectanteGenerico
    template_name = "administrador/genericas/list_desinfectante_generico.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'desinfectante'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la pablabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = DesinfectanteGenerico.objects.filter(
           des_nombre__icontains = palabra_clave
        )
        return lista
    
class DesinfectanteGenericoCreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea una nueva materia prima'''
    model = DesinfectanteGenerico
    template_name = "administrador/genericas/add_desinfectante_generico.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = DesinfectanteGenericoForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('mp_app:desinfeccion_generico')

class DesinfectanteGenericoUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de materia prima generica'''
    model = DesinfectanteGenerico
    template_name = "administrador/genericas/update_desinfectante_generico.html"
    login_url=reverse_lazy('users_app:login')
    form_class=DesinfectanteGenericoForm
    success_url= reverse_lazy('mp_app:desinfeccion_generico')

class DesinfectanteGenericoDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Implenentos de Trabajo'''
    model = DesinfectanteGenerico
    template_name = "administrador/genericas/delete_desinfectante_generico.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('mp_app:desinfeccion_generico')

class DesinfeccionMateriaPrimaCreateView(LoginRequiredMixin, CreateView):
    '''Vists para la creacion de la desinfeccion de la materia prima'''
    model = Desinfeccion
    template_name = "materiaprima/desinfeccion_mp.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = DesinfeccionMPForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('mp_app:lista_mp')
    
    def form_valid(self, form):
        #Obtener los datos del fomulario
        mp_lote = form.cleaned_data['mp_lote']
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        desinfeccion = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        desinfeccion.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        desinfeccion.save()

        # Agregar un mensaje de éxito con el nombre de la materia prima
        messages.success(self.request, f'¡La desinfección de {mp_lote} se ha guardado correctamente!')

        return super(DesinfeccionMateriaPrimaCreateView, self).form_valid(form)
    
class DesinfeccionMateriaPrimaUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para la edición de la desinfeccion de la materia prima'''
    model = Desinfeccion
    template_name = "materiaprima/updateDesinfeccion_mp.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = DesinfeccionMPUpdateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('mp_app:lista_mp')

    def form_valid(self, form):
        #Obtener los datos del fomulario
        mp_lote = form.cleaned_data['mp_lote']
        user = self.request.user
             # Guarda el formulario sin commit para asignar manualmente el usuario
        desinfeccion = form.save(commit=False)
             # Asigna el usuario al campo pedi_user
        desinfeccion.responsable = user
             # Ahora sí, guarda el pedido en la base de datos
        desinfeccion.save()

        # Agregar un mensaje de éxito con el nombre de la materia prima
        messages.success(self.request, f'¡La desinfección de {mp_lote} se ha actualizado correctamente!')

        return super(DesinfeccionMateriaPrimaUpdateView, self).form_valid(form)
        
class MateriaPrimaDetailView(LoginRequiredMixin, DetailView):
    '''Vista donde se muestran los detalles de la materia prima'''
    model = MateriaPrima
    template_name = "materiaprima/detail_mp.html"
    login_url=reverse_lazy('users_app:login')
    context_object_name = 'materiaprima'

class MateriaAuditListView(LoginRequiredMixin, ListView):
    model= MateriaPrimaAudit
    template_name='administrador/auditorias/materiaaudit.html'
    paginate_by=10
    context_object_name='auditoria'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = MateriaAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            materiaprima = form.cleaned_data.get('materiaprima')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if materiaprima:
                queryset = queryset.filter(materiaprima=materiaprima)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = MateriaAuditFilterForm(self.request.GET)
        return context

def export_materiaprima_to_excel(request):
    '''Vista para exportar datos de tabla materia prima en formato excel'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de Materia prima que quieres exportar
    materiaprima = MateriaPrima.objects.all()

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Materia Prima'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    title_alignment = Alignment(horizontal='center')

    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])
    worksheet.merge_cells('A1:R1')
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])

    #blanco entre la información adicional y los encabezados
    worksheet.append([])

    # Agregar fila de título de características organolépticas
    worksheet.merge_cells('G5:K5')  # Fusionar celdas para el título
    worksheet['G5'] = 'CARACTERISTICAS ORGANOLEPTICAS'  # Escribir el título en la celda fusionada

    # Establecer estilos para el título de características organolépticas
    title_font = Font(bold=True, color="000000")  # Letra negra
    title_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro
    title_alignment = Alignment(horizontal='center')

    # Aplicar estilos al título
    title_cell = worksheet['G5']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    
    # Agregar fila de título de características organolépticas
    worksheet.merge_cells('l5:R5')  # Fusionar celdas para el título
    worksheet['L5'] = 'DESINFECCIÓN'  # Escribir el título en la celda fusionada

    # Establecer estilos para el título de características organolépticas
    title_font = Font(bold=True, color="000000")  # Letra negra
    title_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro
    title_alignment = Alignment(horizontal='center')

    # Aplicar estilos al título
    title_cell = worksheet['L5']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    

    # Agregar encabezados a la siguiente fila
    headers = [
        'Lote', 'Materia prima', 'Tipo','Cantidad', 'Fecha de llegada', 'Fecha de vencimiento',
         'Olor', 'Textura','Limpieza','Empaque','Color', 'Estado',
        'Agente desinfectante', 'Concentración','Responsable', 
        'Tiempo de inicio', 'Tiempo de final','Observaciones'
    ]
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de productos terminados a las siguientes filas
    for materia in materiaprima:
        # Obtener las características organolépticas
        try:
            caracteristicas = CaracteristicasOrganolepticas.objects.get(mp_lote=materia)
        except CaracteristicasOrganolepticas.DoesNotExist:
            caracteristicas = None

        # Obtener los datos de desinfeccion
        try:
            desinfeccion = Desinfeccion.objects.get(mp_lote=materia)
        except Desinfeccion.DoesNotExist:
            desinfeccion = None

        data_row = [
            materia.mp_lote,
            materia.mp_nombre.mp_nombre,
            materia.mp_nombre.get_mp_tipo_display(),
            materia.mp_cantidad,
            materia.mp_fechallegada.strftime("%Y-%m-%d"),
            materia.mp_fechavencimiento.strftime("%Y-%m-%d"),
            'Sí' if caracteristicas and caracteristicas.olor else 'No',
            'Sí' if caracteristicas and caracteristicas.textura else 'No',
            'Sí' if caracteristicas and caracteristicas.limpieza else 'No',
            'Sí' if caracteristicas and caracteristicas.empaque else 'No',
            'Sí' if caracteristicas and caracteristicas.color else 'No',
            dict(CaracteristicasOrganolepticas.ESTADO_CHOICES).get(caracteristicas.estado, '') if caracteristicas else '',
            desinfeccion.des_nombre,
            desinfeccion.concentracion,
            desinfeccion.responsable.name,
            desinfeccion.tiempo_inicio.strftime("%Y-%m-%d %H:%M:%S"),
            desinfeccion.tiempo_fin.strftime("%Y-%m-%d %H:%M:%S"),
            desinfeccion.obsevacion,

        ]
        worksheet.append(data_row)

    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=materiaprima.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

def export_materiaprima_to_csv(request):
    '''Vista para exportar datos de tabla producto terminado en formato CSV'''
    materiaprima = MateriaPrima.objects.all()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=materiaprima.csv'

    writer = csv.writer(response)
    headers = [
        'Lote', 'Materia prima', 'Tipo','Cantidad', 'Fecha de llegada', 'Fecha de vencimiento',
         'Olor', 'Textura','Limpieza','Empaque','Color', 'Estado',
        'Agente desinfectante', 'Concentración', 'Responsable'
        'Tiempo de inicio', 'Tiempo de final','Observaciones'
    ]
    writer.writerow(headers)

    for materia in materiaprima:
        # Obtener las características organolépticas
        try:
            caracteristicas = CaracteristicasOrganolepticas.objects.get(mp_lote=materia)
        except CaracteristicasOrganolepticas.DoesNotExist:
            caracteristicas = None

        # Obtener los datos de desinfeccion
        try:
            desinfeccion = Desinfeccion.objects.get(mp_lote=materia)
        except Desinfeccion.DoesNotExist:
            desinfeccion = None

        writer.writerow([
            materia.mp_lote,
            materia.mp_nombre.mp_nombre,
            materia.mp_nombre.get_mp_tipo_display(),
            materia.mp_cantidad,
            materia.mp_fechallegada.strftime("%Y-%m-%d"),
            materia.mp_fechavencimiento.strftime("%Y-%m-%d"),
            'Sí' if caracteristicas and caracteristicas.olor else 'No',
            'Sí' if caracteristicas and caracteristicas.textura else 'No',
            'Sí' if caracteristicas and caracteristicas.limpieza else 'No',
            'Sí' if caracteristicas and caracteristicas.empaque else 'No',
            'Sí' if caracteristicas and caracteristicas.color else 'No',
            dict(CaracteristicasOrganolepticas.ESTADO_CHOICES).get(caracteristicas.estado, '') if caracteristicas else '',
            desinfeccion.des_nombre.des_nombre,
            desinfeccion.concentracion,
            desinfeccion.responsable.name,
            desinfeccion.tiempo_inicio.strftime("%Y-%m-%d %H:%M:%S"),
            desinfeccion.tiempo_fin.strftime("%Y-%m-%d %H:%M:%S"),
            desinfeccion.obsevacion,
            ])
        return response

