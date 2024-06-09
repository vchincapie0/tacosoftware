# Fecha de Creación: 04/02/2024
# Autor: Vivian Carolina Hincapie Escobar
# Última modificación: 15/05/2024

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
# Importación de modelos y formularios
from .models import Proveedores, ProveedoresAudit
from .forms import ProveedorCreateForm, ProveedoresUpdateForm, ProveedorAuditFilterForm

class ProveedoresListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los proveedores'''
    model = Proveedores
    template_name = "proveedores/list_proveedor.html"
    login_url = reverse_lazy('users_app:login')
    paginate_by = 10
    context_object_name = 'proveedor'

    def get_queryset(self):
        '''Función que toma de la barra de búsqueda la palabra clave para filtrar'''
        palabra_clave = self.request.GET.get("kword", '')
        lista = Proveedores.objects.filter(
            prov_nombre__icontains=palabra_clave,
            deleted=False  # Solo proveedores activos
        ).order_by('prov_nombre')  # Ordenar por 'prov_nombre'
        return lista
    
class ProveedoresCreateView(LoginRequiredMixin, CreateView):
    '''Clase para crear nuevos proveedores'''
    model = Proveedores
    template_name = "proveedores/add_proveedor.html"
    login_url = reverse_lazy('home_app:home')
    form_class = ProveedorCreateForm
    success_url = reverse_lazy('proveedores_app:list_proveedores')

    def form_valid(self, form):
        '''Validación del formulario para crear un nuevo proveedor'''
        nombre = form.cleaned_data['prov_nombre']
        messages.success(self.request, f'¡El proveedor {nombre} se ha agregado correctamente!')
        return super(ProveedoresCreateView, self).form_valid(form)

class ProveedorUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de proveedores'''
    model = Proveedores
    template_name = "proveedores/edit_proveedor.html"
    login_url = reverse_lazy('users_app:login')
    form_class = ProveedoresUpdateForm
    success_url = reverse_lazy('proveedores_app:list_proveedores')

    def form_valid(self, form):
        '''Validación del formulario para actualizar un proveedor'''
        nombre = form.cleaned_data['prov_nombre']
        messages.success(self.request, f'¡El proveedor {nombre} se ha actualizado correctamente!')
        return super(ProveedorUpdateView, self).form_valid(form)

class ProveedoresDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar proveedores'''
    model = Proveedores
    template_name = "proveedores/delete_proveedor.html"
    login_url = reverse_lazy('users_app:login')
    success_url = reverse_lazy('proveedores_app:list_proveedores')

class ProveedoresAuditListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de logs de proveedores'''
    model = ProveedoresAudit
    template_name = "administrador/auditorias/proveedoraudit.html"
    login_url = reverse_lazy('users_app:login')
    paginate_by = 10
    context_object_name = 'proveedor'

    def get_queryset(self):
        '''Obtiene y filtra el conjunto de registros de auditoría'''
        queryset = super().get_queryset()
        form = ProveedorAuditFilterForm(self.request.GET)

        if form.is_valid():
            proveedor = form.cleaned_data.get('proveedor')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            if proveedor:
                queryset = queryset.filter(proveedor=proveedor)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        '''Agrega el formulario de filtro al contexto'''
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProveedorAuditFilterForm(self.request.GET)
        return context
    
def export_proveedores_to_excel(request):
    '''Vista para exportar datos de la tabla proveedores en formato Excel'''
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
    proveedores = Proveedores.objects.filter(deleted=False)  # Filtrar proveedores activos

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Proveedores'

    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Transparente
    title_alignment = Alignment(horizontal='center')
    
    worksheet.append(['TACO MAS'])
    worksheet.merge_cells('A1:C1')
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])
    worksheet.append([])  # Agregar una fila vacía

    headers = ['NIT', 'Razón Social', 'Teléfono']
    worksheet.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gris claro
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    for proveedor in proveedores:
        data_row = [proveedor.nit, proveedor.prov_nombre, proveedor.prov_telefono]
        worksheet.append(data_row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    workbook.save(response)

    return response

def export_proveedores_to_csv(request):
    '''Vista para exportar datos de la tabla proveedores en formato CSV'''
    proveedores = Proveedores.objects.filter(deleted=False)  # Obtener datos de proveedores
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=proveedores.csv'

    writer = csv.writer(response)
    writer.writerow(['NIT', 'Razón Social', 'Teléfono'])  # Encabezados de columnas

    for proveedor in proveedores:
        writer.writerow([proveedor.nit, proveedor.prov_nombre, proveedor.prov_telefono])

    return response
