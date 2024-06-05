# Fecha de Creación: 27/02/2024
# Última modificación: 22/05/2024

from django.http import HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.views.generic import (
    ListView, 
    CreateView, 
    UpdateView, 
    DeleteView,
    )
from django.urls import reverse_lazy

#Importacion de modelos y formularios
from .models import Facturas,IVA,FacturasAudit 
from .forms import (
    FacturaCreateForm, 
    FacturaUpdateForm, 
    IVACreateForm, 
    IVAUpdateForm,
    FacturasAuditFilterForm
)

#Autor: Kevin Dayann Albarracin Navarrete
class IVAListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de los Implementos de trabajo'''
    model = IVA
    template_name = "administrador/genericas/iva/list_IVA.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'facturas'

#Autor: Kevin Dayann Albarracin Navarrete
class IVACreateView(LoginRequiredMixin, CreateView):
    '''Clase donde se crea una nueva factura'''
    model = IVA
    template_name = "administrador/genericas/iva/add_IVA.html"
    login_url=reverse_lazy('users_app:login')
    #Campos que se van a mostrar en el formulario
    form_class = IVACreateForm
    #url donde se redirecciona una vez acaba el proceso el "." es para redireccionar a la misma pagina
    success_url= reverse_lazy('facturas_app:list_IVA') 

#Autor: Kevin Dayann Albarracin Navarrete
class IVAUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de Facturas'''
    model =IVA 
    template_name = "administrador/genericas/iva/update_IVA.html"
    login_url=reverse_lazy('users_app:login')
    form_class=IVAUpdateForm
    success_url= reverse_lazy('facturas_app:list_IVA')

#Autor: Kevin Dayann Albarracin Navarrete
class IVADeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar Implenentos de Trabajo'''
    model = IVA
    template_name = "administrador/genericas/iva/delete_IVA.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('facturas_app:list_IVA') 

#Autor:
class FacturasListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de las Facturas'''
    model = Facturas
    template_name = "facturas/list_factura.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'facturas'

    def get_queryset(self):
        '''Funcion que toma de la barra de busqueda la pablabra clave para filtrar'''
        palabra_clave= self.request.GET.get("kword",'')
        lista = Facturas.objects.filter(
            num_factura__icontains = palabra_clave,
            deleted=False  # Solo usuarios activos
        )
        return lista

#Autor:
class FacturasCreateView(LoginRequiredMixin, CreateView):
    '''Clase encargada de vista de crear factura'''
    model = Facturas
    template_name = "facturas/add_fact.html"
    login_url = reverse_lazy('users_app:login')
    form_class = FacturaCreateForm
    success_url = reverse_lazy('facturas_app:list_factura')  

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        numeroFactura = form.cleaned_data['num_factura']
        pedido = form.cleaned_data['fac_numeroPedido']

        # Verificar si ya existe una factura para el mismo pedido
        if Facturas.objects.filter(fac_numeroPedido=pedido).exists():
            form.add_error('fac_numeroPedido', 'Ya existe una factura para este pedido.')
            messages.error(self.request, 'Ya existe una factura para el pedido seleccionado.')
            return self.form_invalid(form)

        # Agregar un mensaje de éxito con el número de factura
        messages.success(self.request, f'¡La factura {numeroFactura} se ha agregado correctamente!')
        return super(FacturasCreateView, self).form_valid(form)
#Autor:
class FacturasUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos de Facturas'''
    model =Facturas 
    template_name = "facturas/edit_factura.html"
    login_url=reverse_lazy('users_app:login')
    form_class=FacturaUpdateForm
    success_url= reverse_lazy('facturas_app:list_factura')

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def form_valid(self, form):
        numeroFactura = form.cleaned_data['num_factura']

        # Agregar un mensaje de éxito con el numero de factura
        messages.success(self.request, f'¡La factura {numeroFactura} se ha actualizado correctamente!')

        return super(FacturasUpdateView, self).form_valid(form)

#Autor:
class FacturasDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar facturas'''
    model = Facturas
    template_name = "facturas/delete_fact.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('facturas_app:list_factura')

#Autor:Vivian Carolina Hincapie Escobar
class FacturasAuditListView(LoginRequiredMixin, ListView):
    '''Clase para mostrar los datos de las auditorias de Facturas'''
    model = FacturasAudit
    template_name = "administrador/auditorias/factura_audit.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=10
    context_object_name = 'facturas'

    def get_queryset(self):
        ''''Filtro de la vista'''
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = FacturasAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            factura = form.cleaned_data.get('factura')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            pedido=form.cleaned_data.get('pedido')
            proveedor=form.cleaned_data.get('proveedor')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if factura:
                queryset = queryset.filter(factura=factura)
            if pedido:
                queryset = queryset.filter(pedido=pedido)
            if proveedor:
                queryset = queryset.filter(proveedor=proveedor)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = FacturasAuditFilterForm(self.request.GET)
        return context
    
#Autor:Vivian Carolina Hincapie Escobar   
def export_facturas_to_excel(request):
    '''Vista de exportación en archivo tipo excel de los datos en el modelo Facturas'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de proveedores que quieres exportar
    facturas = Facturas.objects.filter(deleted=False)  # Filtrar proveedores activos

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Facturas'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Transparente
    title_alignment = Alignment(horizontal='left')
    
    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])  # Agregar texto del título
    worksheet.merge_cells('A1:H1')  # Combinar celdas para el título
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])

    # Agregar espacio en blanco entre la información adicional y los encabezados
    worksheet.append([])  # Agregar una fila vacía

    # Agregar encabezados a la siguiente fila
    headers = ['Número Factura', 'Proveedor', 'Pedido', 'Fecha de Llegada','Unidades', 'Subtotal','IVA','Total']
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gris claro
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de factura a las siguientes filas
    # Iterar sobre cada factura y agregar sus datos a la hoja de trabajo
    for factura in facturas:
        # Obtener los valores necesarios de los modelos relacionados
        proveedor_nombre = factura.fac_proveedor.prov_nombre if factura.fac_proveedor else ''
        numero_pedido = factura.fac_numeroPedido.ref_pedido if factura.fac_numeroPedido else ''
        iva_valor = factura.fac_iva.valor if factura.fac_iva else 0  # Obtener el valor del IVA

        # Construir la fila de datos para la factura
        data_row = [
            factura.num_factura,
            proveedor_nombre,
            numero_pedido,
            factura.fac_fechaLlegada,
            factura.fac_numeroUnidades,
            factura.fac_subtotal,
            iva_valor,
            factura.fac_total,
        ]

        # Agregar la fila de datos a la hoja de trabajo
        worksheet.append(data_row)

    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=facturas.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

#Autor:Vivian Carolina Hincapie Escobar
def export_facturas_to_csv(request):
    '''Vista de exportación en archivo tipo CSV de los datos en el modelo Facturas'''
    
    facturas = Facturas.objects.filter(deleted=False)  # Obtener datos de facturas
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=facturas.csv'

    writer = csv.writer(response)
    writer.writerow(['Número Factura', 'Proveedor', 'Pedido', 'Fecha de Llegada','Unidades', 'Subtotal','IVA','Total'])  # Encabezados de columnas

    for factura in facturas:
        # Obtener los valores necesarios de los modelos relacionados
        proveedor_nombre = factura.fac_proveedor.prov_nombre if factura.fac_proveedor else ''
        numero_pedido = factura.fac_numeroPedido.ref_pedido if factura.fac_numeroPedido else ''
        iva_valor = factura.fac_iva.valor if factura.fac_iva else 0  # Obtener el valor del IVA

        # Construir la fila de datos para la factura
        writer.writerow([
            factura.num_factura,
            proveedor_nombre,
            numero_pedido,
            factura.fac_fechaLlegada,
            factura.fac_numeroUnidades,
            factura.fac_subtotal,
            iva_valor,
            factura.fac_total,
        ]) 

    return response