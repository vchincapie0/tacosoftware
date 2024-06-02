#Fecha de Creación: 02/02/2024
#Autor: Vivian Carolina Hincapie Escobar
#Última modficación: 15/05/2024

#Importación de Librerias
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from django.http import HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import  redirect
from django.views.generic import View,ListView,UpdateView, FormView, DeleteView
from django.contrib.auth.views import LoginView
from django.views.generic.edit import FormView
from django.urls import reverse_lazy,reverse

#Importacion modelos y formularios
from .models import User, UserAudit
from .forms import UserRegisterForm, UserUpdateForm, UserAuditFilterForm

class UsersListView(LoginRequiredMixin, ListView):
    '''vista que muestra todos los usuarios registrados'''
    model = User
    template_name = "usuarios/list_user.html"
    login_url=reverse_lazy('users_app:login')
    paginate_by=5
    context_object_name='usuarios'

    def get_queryset(self):
        palabra_clave= self.request.GET.get("kword",'')
        lista = User.objects.filter(
            name__icontains = palabra_clave,
            deleted=False # Solo usuarios activos
        )
        return lista

class UserRegisterView(LoginRequiredMixin,FormView):
    '''Vista que registra usuarios del modelo user'''
    template_name ='usuarios/add_user.html'
    login_url=reverse_lazy('users_app:login')
    form_class=UserRegisterForm
    success_url=reverse_lazy('users_app:list_user')

    def form_valid(self, form):
        '''Función para guardar los datos del usuario'''
        # Obtener los datos del formulario
        username = form.cleaned_data['username']
        name = form.cleaned_data['name']
        last_name = form.cleaned_data['last_name']
        password = form.cleaned_data['password']
        is_admin = form.cleaned_data['is_admin']

        # Crear el usuario en la base de datos
        User.objects.create_user(
            username=username,
            name=name,
            last_name=last_name,
            password=password,
            is_admin=is_admin,
        )

        # Agregar un mensaje de éxito con el nombre de usuario
        messages.success(self.request, f'¡El usuario {username} se ha agregado correctamente!')

        return super(UserRegisterView, self).form_valid(form)
    
class UserUpdateView(LoginRequiredMixin, UpdateView):
    '''Vista para actualizar los datos del usuario'''
    model = User
    template_name = "usuarios/update_user.html"
    login_url=reverse_lazy('users_app:login')
    form_class=UserUpdateForm
    success_url= reverse_lazy('users_app:list_user')

    def get_object(self, queryset=None):
        # Obtener el usuario basado en el parámetro pk de la URL
        pk = self.kwargs.get('pk')
        return get_object_or_404(User, pk=pk)

    def form_valid(self, form):
        # # Obtener el usuario actualizado desde el formulario
        user = form.instance
        username = user.username
        new_password = form.cleaned_data.get('password')

        if new_password:
            #Si se proporciona una nueva contraseña, encriptarla y guardarla
            user.set_password(new_password)

            # Guardar el usuario actualizado
            user.save()

        messages.success(self.request, f'¡Los cambios de {username} se han guardado correctamente!')

        return super().form_valid(form)

class UserDeleteView(LoginRequiredMixin, DeleteView):
    '''Vista para borrar usuarios'''
    model = User
    template_name = "usuarios/delete_user.html"
    login_url=reverse_lazy('users_app:login')
    success_url= reverse_lazy('users_app:list_user')

class LogIn(LoginView):
    '''Vista para login'''
    template_name = 'usuarios/login.html'
    success_url = reverse_lazy('home_app:home')

    def form_valid(self, form):
        # Recoge los datos del formulario
        username = form.cleaned_data['username']
        password = form.cleaned_data['password']

        # Autentica al usuario utilizando el sistema de autenticación de Django
        user = authenticate(username=username, password=password)

        if user is not None:
            # Inicia sesión para el usuario autenticado
            login(self.request, user)
            return redirect(self.get_success_url())

class LogOut(View):
    '''Vista para cerrar sesión'''
    def get(self, request, *args, **kwargs):
        logout(request)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # Si la solicitud tiene el encabezado 'x-requested-with' establecido en 'XMLHttpRequest',
            # entonces es una solicitud AJAX, responde con un mensaje JSON
            return JsonResponse({'message': 'Sesión cerrada correctamente.'})
        else:
            # Si no es una solicitud AJAX, redirige a la página de inicio de sesión
            return HttpResponseRedirect(reverse('users_app:login'))
    
class UserAuditListView(LoginRequiredMixin, ListView):
    '''Vista para la lista de auditorias de usuarios'''
    model= UserAudit
    template_name='administrador/auditorias/useraudit.html'
    paginate_by=10
    context_object_name='auditoria'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los parámetros de filtrado del formulario
        form = UserAuditFilterForm(self.request.GET)

        # Aplicar filtros si el formulario es válido
        if form.is_valid():
            user = form.cleaned_data.get('user')
            action = form.cleaned_data.get('action')
            changed_by = form.cleaned_data.get('changed_by')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            # Filtrar por usuario, acción, usuario que realizó el cambio y rango de fechas
            if user:
                queryset = queryset.filter(user=user)
            if action:
                queryset = queryset.filter(action=action)
            if changed_by:
                queryset = queryset.filter(changed_by=changed_by)
            if start_date:
                queryset = queryset.filter(changed_at__gte=start_date)
            if end_date:
                # Agregar 1 día a la fecha final para incluir todos los registros de ese día
                end_date += timezone.timedelta(days=1)
                queryset = queryset.filter(changed_at__lt=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = UserAuditFilterForm(self.request.GET)
        return context
    
def export_users_to_excel(request):
    '''Vista para exportar datos de tabla Usuarios en formato excel'''
    # Obtener la fecha y hora actual
    fecha_descarga = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # Obtener los datos de proveedores que quieres exportar
    usuarios =User.objects.filter(deleted=False)  # Filtrar proveedores activos

    # Crear un nuevo libro de Excel y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Usuarios'

    # Establecer estilos para la primera línea (encabezado personalizado)
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Transparente
    title_alignment = Alignment(horizontal='center')
    
    # Agregar fila de título personalizado
    worksheet.append(['TACO MAS'])  # Agregar texto del título
    worksheet.merge_cells('A1:D1')  # Combinar celdas para el título
    title_cell = worksheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    # Agregar información adicional (fecha y nombre del software) en una nueva fila
    worksheet.append(['Fecha de descarga:', fecha_descarga])
    worksheet.append(['Software:', 'Tacosoft'])
    worksheet.append(['Tabla de Registro:', 'Usuarios Registrados'])


    # Agregar espacio en blanco entre la información adicional y los encabezados
    worksheet.append([])  # Agregar una fila vacía

    # Agregar encabezados a la siguiente fila
    headers = ['Usuario', 'Nombre', 'Apellido', 'Rol']
    worksheet.append(headers)

    # Aplicar estilos a la fila de encabezados (fila actual + 2)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gris claro
    header_alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        header_cell = worksheet.cell(row=worksheet.max_row, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Agregar datos de usuarios a las siguientes filas
    for usuario in usuarios:
        # Determinar el rol del usuario
        if usuario.is_admin:
            rol = 'Administrador'
        else:
            rol = 'Operario'

        data_row = [usuario.username, usuario.name, usuario.last_name,rol]
        worksheet.append(data_row)

    # Crear una respuesta HTTP con el archivo Excel como contenido
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    workbook.save(response)

    return response

def export_users_to_csv(request):
    '''Vista para exportar datos de tabla usuarios en formato CSV'''
    usuarios = User.objects.filter(deleted=False)  # Obtener datos de usuarios
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=usuarios.csv'

    writer = csv.writer(response)
    writer.writerow(['Usuario', 'Nombre', 'Apellido', 'Rol'])  # Encabezados de columnas

    for usuario in usuarios:
        # Determinar el rol del usuario
        if usuario.is_admin:
            rol = 'Administrador'
        else:
            rol = 'Operario'

        writer.writerow([usuario.username, usuario.username, usuario.name, usuario.last_name, rol])

    return response